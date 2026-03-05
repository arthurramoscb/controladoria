# Databricks notebook source
# MAGIC %md
# MAGIC ## Modelo Provisão - Prepara a base do Fechamento Financeiro
# MAGIC
# MAGIC **Inputs**\
# MAGIC Planilha Fechamento Financeiro (ff) - Prévia ou Fechamento\
# MAGIC Base Gerencial Trabalhista
# MAGIC
# MAGIC **Outputs**\
# MAGIC Base Fechamento Finaceiro tratado tb_fechamento_trabalhista

# COMMAND ----------

# MAGIC %md
# MAGIC #####Define e cria todos os _Widgets_ utilizados na atualização das bases

# COMMAND ----------

# Configura campo para que o usuário insira parâmetros

# Parametro do formato de data do arquivo financeiro. Ex: 22.02.2024
dbutils.widgets.text("nmtabela_finan", "")

# Parametro de data do fechamento do eLaw
dbutils.widgets.text("fech_elaw", "2024-05-16", "Selecione uma data")
# dbutils.widgets.remove("fech_elaw")

# Parametro de data da tabela trab_ger_consolidado_ . Ex: 20240423
dbutils.widgets.text("nmtabela_trab_ger_consolidado", "")

# Parametro de data da tabela MESFECH. Ex: 01/04/2024
# Será utilizado como valor de uma coluna no join com o passo anterior
dbutils.widgets.text("mes_fechamento", "")

# Parametro para escolher tipo do arquivo. Ex: prévia ou fechamento
dbutils.widgets.dropdown("tipo_arquivo", "Fechamento", ["Prévia","Fechamento"])

# COMMAND ----------

# MAGIC %md
# MAGIC #####Carrega as funções para o tratamento das bases

# COMMAND ----------

# MAGIC %md
# MAGIC #####Lista todos os arquivos do diretório especificado abaixo

# COMMAND ----------

# Listar todos os arquivos dentro do diretório especificado
from pyspark.sql.functions import *

# Caminho do diretório
dir_path = '/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/'

# Usar dbutils para listar os arquivos
files_list = dbutils.fs.ls(dir_path)

# Criar um DataFrame com os nomes dos arquivos
df_files = spark.createDataFrame(files_list).select("name")

# Adicionar uma coluna com o caminho completo do arquivo
df_files_with_path = df_files.withColumn("full_path", concat(lit(dir_path),df_files['name']))

# Carregar os valores da coluna em uma lista
coluna_lista = df_files_with_path.select("name").collect()

# Exibir o DataFrame
display(df_files_with_path)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Altera o nome do arquivo

# COMMAND ----------

import pandas as pd
from openpyxl import load_workbook
from pyspark.sql.functions import col, concat, lit

# Obter o nome da parte do arquivo para filtrar
tipo_arquivo = dbutils.widgets.get("tipo_arquivo")
nmtabela_finan = dbutils.widgets.get("nmtabela_finan")
parte_nome_arquivo = nmtabela_finan

# Caminho do diretório
dir_path = '/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/'

# Listar os arquivos no diretório
files_list = dbutils.fs.ls(dir_path)

# Criar um DataFrame com os nomes dos arquivos
df_files = spark.createDataFrame(files_list).select("name")

# Filtrar arquivos que contêm a parte do nome especificada
df_filtered_files = df_files.filter(col("name").contains(parte_nome_arquivo))

# Adicionar uma coluna com o caminho completo do arquivo
df_filtered_files_with_path = df_filtered_files.withColumn("full_path", concat(lit(dir_path), lit("/"), col("name")))

# Coletar os valores da coluna 'name' em uma lista de objetos Row
nome_arquivos = df_filtered_files_with_path.select("name").collect()

# Exibir o DataFrame com os arquivos filtrados e seus caminhos completos
display(df_filtered_files_with_path)
print(nome_arquivos)

# Verificar se o nome do arquivo atende a regra de negócio e renomeá-lo se necessário
file_name = nome_arquivos[0]["name"] 
dir_path = "/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/"
full_file_path = dir_path + file_name

# Nomes de arquivo válidos de acordo com a regra de negócio
valid_names = [f"{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx"]
# Verificar se o nome do arquivo não está na lista de nomes válidos
if file_name not in valid_names:
    # Definir o novo nome do arquivo aqui. Exemplo: adicionando " Renomeado" ao nome original
    new_file_name = file_name.replace(file_name, valid_names[0])
    new_file_path = dir_path + new_file_name
    
    # Renomear o arquivo
    dbutils.fs.mv(full_file_path, new_file_path)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Cria as _Paths´s_ com os diretórios usados na importação dos arquivos

# COMMAND ----------

# Caminho das pastas e arquivos
nmtabela_finan = dbutils.widgets.get("nmtabela_finan")
tipo_arquivo = dbutils.widgets.get("tipo_arquivo")

# Carrega o diretório da base financeira (prévia/fechamento)
path_ff = f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx'

# Carrega o diretório da base auxiliar com o de para de assuntos
path_dp_assunto = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_auxiliares/DE_PARA_ASSUNTO_CARGO.xlsx'

# Carrega o diretório da base auxiliar com o de para de cluster de valor
path_dp_cluster_valor = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_auxiliares/DE_PARA_CLUSTER_VALOR.xlsx'

# Carrega o diretório da base auxiliar com o de para de comarca
path_dp_comarca = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_auxiliares/DE_PARA_COMARCA.xlsx'

# Carrega o diretório da base auxiliar com o de para de terceiro insolvente
path_dp_terceiro = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_auxiliares/DE_PARA_TERCEIRO_INSOLVENTE_V1.xlsx'

# COMMAND ----------

# MAGIC %md
# MAGIC #####Lista todas as _Sheet´s_ dentro do arquivo da base de fechamento

# COMMAND ----------

# Identifica e lista as planilhas do arquivo excel

# Import the required libraries
import pandas as pd

# Read the excel file
df_dict_ff = pd.read_excel(path_ff, sheet_name=None)

# Access the DataFrames of each sheet
for sheet_name, df_fech_fin in df_dict_ff.items():
    print(f"Nome da planilha: {sheet_name}")
    # Perform the desired processing on the DataFrames
    # For example, show the first 5 rows
    #print(df_dp_assunto.head(5))

# COMMAND ----------

# MAGIC %md
# MAGIC #####Altera o nome da planilha para o padrão definido

# COMMAND ----------

import pandas as pd
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from shutil import copyfile

nmtabela_finan = dbutils.widgets.get("nmtabela_finan")
tipo_arquivo = dbutils.widgets.get("tipo_arquivo")

# Assuming nmtabela_finan and tipo_arquivo are defined somewhere in your code

# Load the workbook with openpyxl
arquivo_excel = f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx'

workbook = load_workbook(arquivo_excel)

# Iterate over the sheets and rename as necessary
for sheet in workbook.sheetnames:
    if sheet == "FINAL":
        workbook[sheet].title = "Base Fechamento"
    elif sheet == "Base Fechamento":
        pass  # Do nothing, keep the name "Base Fechamento"


# Salva o workbook em um arquivo temporário
with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    workbook.save(tmp.name)
    temp_filename = tmp.name

# Save the DataFrame with the corrected names to an Excel file on the local disk first
local_path = f'{temp_filename}'

# Copy the file from the local disk to the desired volume
volume_path = f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx'

copyfile(local_path, volume_path)

print(f'Workbook salvo em {temp_filename}')

# COMMAND ----------

# MAGIC %md
# MAGIC #####Corrige as colunas com os mesmos nomes e salva o arquivo

# COMMAND ----------

# Testes jefferson

import pandas as pd
import re
from shutil import copyfile
from pyspark.sql import SparkSession
from pyspark.sql.functions import udf, col, regexp_replace
from pyspark.sql.types import StringType
import unidecode

# Aqui não usei o header=3 por que o arquivo já esta tratado, porém usei usecols='A:BQ'
# Se o range for padrão, o parametro de colunas selecionadas pode ser passado na leitura do df pelo pandas

df = pd.read_excel(f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx', sheet_name='Base Fechamento', usecols='A:BQ')

print(len(df))

display(df)

# COMMAND ----------

import pandas as pd
import re
from shutil import copyfile

nmtabela_finan = dbutils.widgets.get("nmtabela_finan")
tipo_arquivo = dbutils.widgets.get("tipo_arquivo")

# Função para renomear colunas duplicadas
def rename_duplicate_columns(df):
    # Normaliza espaços em branco nas colunas, transforma todas para minúsculas
    cols_normalized = pd.Series(df.columns.str.lower().str.replace(r'\s+', ' ', regex=True))
    
    # Dicionário para armazenar contagens de duplicatas
    counts = {}
    
    # Lista para novos nomes de colunas
    new_columns = []
    
    # Itera sobre as colunas originais
    for col in df.columns:
        col_normalized = col.lower()
        col_normalized = re.sub(r'\s+', ' ', col_normalized)
        if col_normalized in counts:
            counts[col_normalized] += 1
            new_columns.append(f"{col}_{counts[col_normalized]}")
        else:
            counts[col_normalized] = 0
            new_columns.append(col)
    
    # Atualiza os nomes das colunas no DataFrame
    df.columns = new_columns
    return df

# Ler o arquivo Excel
df = pd.read_excel(f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx', sheet_name='Base Fechamento', header=3)

# Renomear colunas duplicadas
df = rename_duplicate_columns(df)

# Função para converter letras de colunas para índices numéricos
def col_letter_to_index(letter):
    index = 0
    for i, char in enumerate(reversed(letter)):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1

# Definir as letras das colunas de interesse
col_start_letter = 'A'
col_end_letter = 'BQ'

# Converter as letras para índices numéricos
col_start_index = col_letter_to_index(col_start_letter)
col_end_index = col_letter_to_index(col_end_letter)

# Selecionar as colunas usando os índices numéricos
df = df.iloc[:, col_start_index:col_end_index + 1]

# Salvar o DataFrame com os nomes corrigidos em um arquivo Excel no disco local primeiro
local_path = f'/local_disk0/tmp/{nmtabela_finan} {tipo_arquivo} Automacao.xlsx'
df.to_excel(local_path, index=False, sheet_name='Base Fechamento')

# Copiar o arquivo do disco local para o volume desejado
volume_path = f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx'

copyfile(local_path, volume_path)

# Salvar todos os nomes das colunas em uma única coluna de um novo DataFrame
columns_df = pd.DataFrame(df.columns, columns=['Column Names'])

# Exibir o DataFrame com os nomes das colunas
display(columns_df)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Importa a base de fechamento e as bases auxiliares. Cria os Dataframes Pyspark

# COMMAND ----------

# Carrega as planilhas em Spark Data Frames
from pyspark.sql import SparkSession

nmtabela_finan = dbutils.widgets.get("nmtabela_finan")
tipo_arquivo = dbutils.widgets.get("tipo_arquivo")

# Carrega o diretório da base financeira (prévia/fechamento)
path_ff = f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/input/{nmtabela_finan} {tipo_arquivo} Trabalhista Automacao.xlsx'

# base financeira (prévia/fechamento)
df_ff = read_excel(path_ff)

# Converter PySpark DataFrame para Pandas DataFrame
pandas_df = df_ff.toPandas()

# base auxiliar com o de para de assuntos
df_dp_assunto = read_excel(path_dp_assunto, "Objeto_Assunto_Cargo!A1")

# remover espaços em branco no início e no final / substituir múltiplos espaços em branco consecutivos por um único espaço
df_dp_assunto = compress_values(df_dp_assunto, ["Objeto Assunto/Cargo (M)", "Cargo Tratado"])

# base auxiliar com o de para de cluster de valor
df_dp_cluster_valor = read_excel(path_dp_cluster_valor, "Cluster_Valor!A1")

# remover espaços em branco no início e no final / substituir múltiplos espaços em branco consecutivos por um único espaço
df_dp_cluster_valor = compress_values(df_dp_cluster_valor, ["DE", "PARA"])

# base auxiliar com o de para de comarca
df_dp_comarca = read_excel(path_dp_comarca, "DE_PARA_COMARCA!A1")

# remover espaços em branco no início e no final / substituir múltiplos espaços em branco consecutivos por um único espaço
df_dp_comarca = compress_values(df_dp_comarca, ["COMARCA", "DP_COMARCA"])

# base auxiliar com o de para de terceiro insolvente
df_dp_terceiro = read_excel(path_dp_terceiro, "DE_PARA_TI!A1")

# remover espaços em branco no início e no final / substituir múltiplos espaços em branco consecutivos por um único espaço
df_dp_terceiro = compress_values(df_dp_terceiro, ["EMPRESA TERCEIRIZADA", "TERCEIRO_AJUSTADO"])

# display(df_dp_terceiro)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Ajusta os nomes das colunas - substitui os caracteres especiais - exclui as linhas em branco

# COMMAND ----------

from pyspark.sql.functions import col

# Ajusta os nomes das colunas e substitui os caracteres especiais por "_" (Esses caracteres especiais geram muitos erros)
df_ff = adjust_column_names(df_ff)

# Remove as linhas em branco onde as colunas abaixo são nulas
df_ff.dropna(subset=['LINHA', 'ID_PROCESSO', 'ÁREA_DO_DIREITO'])

# Conta a quantidade de linhas do dataframe
num_linhas = df_ff.count()
print(f"Quantidade de linhas: {num_linhas}")

display(df_ff.limit(5))

# COMMAND ----------

# MAGIC %md
# MAGIC ######Exclui as colunas desnecessárias do Dataframe

# COMMAND ----------

from pyspark.sql.functions import col

# Cria a lista que correspondem à parte do nome das colunas que serão excluídas
parte_do_nome = [s.upper() for s in ["Cluster_Aging_Tempo_de_Empresa", "Cluster_Aging", "cluster_valor", "cargo_tratado", "Safra_de_Reclamação"]]

# Filtrar as colunas que correspondem à parte do nome
#colunas_a_excluir = [coluna for coluna in df_ff.columns if 'Cluster_Aging' in coluna]
colunas_a_excluir = [coluna for coluna in df_ff.columns if any(nome in coluna for nome in parte_do_nome)]

# Excluir as colunas filtradas
df_ff = df_ff.drop(*colunas_a_excluir)
display(df_ff.limit(5))

# Conta a quantidade de linhas do dataframe
num_linhas = df_ff.count()
print(f"Quantidade de linhas: {num_linhas}")

# COMMAND ----------

# MAGIC %md
# MAGIC #####Converte as colunas listadas para o formato data

# COMMAND ----------

# Converte as colunas listadas para o tipo data
colunas_data = ['CADASTRO','REABERTURA','DISTRIBUIÇÃO']

df_ff = convert_to_date_format(df_ff, colunas_data)
display(df_ff.limit(5))

# Conta a quantidade de linhas do dataframe
num_linhas = df_ff.count()
print(f"Quantidade de linhas: {num_linhas}")

# COMMAND ----------

# MAGIC %md
# MAGIC #####Converte as colunas listadas para o formato número

# COMMAND ----------

# Converte as colunas listadas para o tipo número
colunas_numeros = ['CENTRO_DE_CUSTO_M_1', 'CENTRO_DE_CUSTO_M']

df_ff = convert_to_float(df_ff, colunas_numeros)
display(df_ff.limit(5))

# Conta a quantidade de linhas do dataframe
num_linhas = df_ff.count()
print(f"Quantidade de linhas: {num_linhas}")

# COMMAND ----------

# MAGIC %md
# MAGIC #####Lista todas as _Sheet´s_ dentro do arquivo da base gerencial

# COMMAND ----------

# Identifica e lista as planilhas do arquivo excel
import pandas as pd

nmtabela_trab_ger = dbutils.widgets.get("nmtabela_trab_ger_consolidado")

path_trab_ger = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_gerenciais/external/TRABALHISTA_GERENCIAL_(CONSOLIDADO)-{nmtabela_trab_ger}.xlsx'

# Read the excel file
df_dict_trab_ger = pd.read_excel(path_trab_ger, sheet_name=None)

# Access the DataFrames of each sheet
for sheet_name, df_trab_ger in df_dict_trab_ger.items():
    print(f"Nome da planilha: {sheet_name}")


# COMMAND ----------

# MAGIC %md
# MAGIC #####Carrega a base gerencial no Pandas e salva no databox

# COMMAND ----------

# Importa a biblioteca pandas
import pandas as pd
import re
from shutil import copyfile

# Carrega as planilhas
nmtabela_trab_ger = dbutils.widgets.get("nmtabela_trab_ger_consolidado")

path_trab_ger = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_gerenciais/external/TRABALHISTA_GERENCIAL_(CONSOLIDADO)-{nmtabela_trab_ger}.xlsx'

df_trab_ger = pd.read_excel(path_trab_ger, sheet_name='TRABALHISTA', header=5)

# Salvar o DataFrame com os nomes corrigidos em um arquivo Excel no disco local primeiro
local_path = f'/local_disk0/tmp/TRABALHISTA_GERENCIAL_(CONSOLIDADO)-{nmtabela_trab_ger}.xlsx'
df_trab_ger.to_excel(local_path, index=False, sheet_name='TRABALHISTA')

# Copiar o arquivo do disco local para o volume desejado
volume_path = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_gerenciais/external/TRABALHISTA_GERENCIAL_(CONSOLIDADO)-{nmtabela_trab_ger}_F.xlsx'

copyfile(local_path, volume_path)


# COMMAND ----------

# MAGIC %md
# MAGIC #####Importa a base gerencial e cria o Dataframe Pyspark - formata campo data

# COMMAND ----------

# Carrega as planilhas em Spark Data Frames
nmtabela_trab_ger = dbutils.widgets.get("nmtabela_trab_ger_consolidado")

path_trab_ger = f'/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_gerenciais/external/TRABALHISTA_GERENCIAL_(CONSOLIDADO)-{nmtabela_trab_ger}_F.xlsx'

df_trab_ger = read_excel(path_trab_ger, "'TRABALHISTA'!A1")

# Ajusta os nomes das colunas e substitui os caracteres especiais por "_" (Esses caracteres especiais geram muitos erros)
df_trab_ger = adjust_column_names(df_trab_ger)

# Converte as colunas listadas para o tipo data
# colunas_data = ['PARTE_CONTRÁRIA_DATA_ADMISSÃO','PARTE_CONTRÁRIA_DATA_DISPENSA']
# df_trab_ger = convert_to_date_format(df_trab_ger, colunas_data)

# # Convertendo as colunas para o formato de data
df_trab_ger = df_trab_ger.withColumn(
    "PARTE_CONTRÁRIA_DATA_ADMISSÃO", 
    to_date(col("PARTE_CONTRÁRIA_DATA_ADMISSÃO"))
).withColumn(
    "PARTE_CONTRÁRIA_DATA_DISPENSA", 
    to_date(col("PARTE_CONTRÁRIA_DATA_DISPENSA"))
)

# remover espaços em branco no início e no final / substituir múltiplos espaços em branco consecutivos por um único espaço
df_trab_ger = compress_values(df_trab_ger, ["TERCEIRO_PRINCIPAL","NOVO_TERCEIRO","PROCESSO_ESTADO"])

display(df_trab_ger)

# COMMAND ----------

df_trab_ger.select('PARTE_CONTRÁRIA_DATA_ADMISSÃO','PARTE_CONTRÁRIA_DATA_DISPENSA').display()

# COMMAND ----------

# MAGIC %md
# MAGIC #####Conta e valida a quantidade de linhas e colunas da base gerencial 

# COMMAND ----------

# Contar a quantidade de linhas
num_linhas = df_trab_ger.count()
print(f"Quantidade de linhas: {num_linhas}")

# Contar a quantidade de colunas
num_colunas = len(df_trab_ger.columns)
print(f"Quantidade de colunas: {num_colunas}")

# COMMAND ----------

# MAGIC %md
# MAGIC #####Lista as colunas da base gerencial para análise

# COMMAND ----------

# Lista e ordena todos os nomes das colunas do DataFrame
from pyspark.sql.functions import lit
from pyspark.sql import Row

# Cria uma lista com os nomes das colunas
column_names = sorted(df_trab_ger.columns)

# Cria Rows onde cada linha contém um nome de coluna
rows = [Row(column_name=column) for column in column_names]

# Cria um DataFrame a partir das Rows com uma coluna "column_name"
df_column_names = spark.createDataFrame(rows)

# Exibir o novo DataFrame
display(df_column_names)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Faz o Join com as bases de fechamento e gerencial:
# MAGIC - carrega as colunas adicionais
# MAGIC - cria a coluna DE_PARA_FASE

# COMMAND ----------

# Dropando campos
df_ff = df_ff.drop("NOVOS") \
               .drop("ENCERRADOS")\
               .drop("ESTOQUE") 
                                           
df_ff.createOrReplaceTempView("FECHAMENTO_TRAB_2")
df_trab_ger.createOrReplaceTempView(f"TRAB_GER_CONSOLIDA")
mes_fechamento = dbutils.widgets.get("mes_fechamento")

df_fechamento_trab = spark.sql(f"""
/* CARREGA AS INFORMAÇÕES ADICIONAIS NA BASE DO FECHAMENTO */
SELECT A.*
	,B.`PARTE_CONTRÁRIA_CPF` AS PARTE_CONTRARIA_CPF
	,1 AS NOVOS
	,1 AS ENCERRADOS
	,1 AS ESTOQUE
	,A.`DISTRIBUIÇÃO`
	/*,MDY(MONTH(A.`DISTRIBUIÇÃO`),01,YEAR(A.`DISTRIBUIÇÃO`)) AS MES_DISTRIBUICAO */
 	,to_date(concat(year(A.`DISTRIBUIÇÃO`), '-', month(A.`DISTRIBUIÇÃO`), '-', '01')) AS MES_DISTRIBUICAO
	,B.DATA_REGISTRADO
	
	,(CASE WHEN A.CADASTRO IS NULL THEN B.DATA_REGISTRADO
		ELSE A.CADASTRO END) AS CADASTRO_AJUSTADO
      
	,(CASE WHEN A.CADASTRO IS NULL THEN to_date(concat(year(B.DATA_REGISTRADO), '-', month(B.DATA_REGISTRADO), '-', '01'))
		ELSE to_date(concat(year(A.CADASTRO), '-', month(A.CADASTRO), '-', '01')) END ) AS MES_CADASTRO

    ,B.FASE

  ,(CASE WHEN B.FASE IN ('' 'N/A' 'INATIVO' 'ENCERRADO' 'ENCERRAMENTO' 'ADMINISTRATIVO') THEN 'DEMAIS'
		WHEN B.FASE IN ('EXECUÇÃO',
					'EXECUÇÃO - INATIVO',
					'EXECUÇÃO - TRT',
					'EXECUÇÃO - TRT - INATIVO',
					'EXECUÇÃO - TST',
	     			'EXECUÇÃO DEFINITIVA',
					'EXECUÇÃO DEFINITIVA (TRT)',
					'EXECUÇÃO DEFINITIVA (TST)',
					'EXECUÇÃO DEFINITIVA PROSSEGUIMENTO',
					'EXECUÇÃO PROVISORIA (TRT)',
					'EXECUÇÃO PROVISORIA (TST)',
					'EXECUÇÃO PROVISÓRIA',
					'EXECUÇÃO PROVISÓRIA - INATIVO',
					'EXECUÇÃO PROVISÓRIA PROSSEGUIMENTO') THEN 'EXECUÇÃO'
		WHEN B.FASE IN ('RECURSAL',
					'RECURSAL - INATIVO',
					'RECURSAL TRT',
					'RECURSAL TRT - INATIVO',
					'RECURSAL TST',
					'RECURSAL TST - INATIVO') THEN 'RECURSAL' 
		ELSE B.FASE END) AS DE_PARA_FASE

	,B.PROCESSO_ESTADO AS ESTADO
	,B.TERCEIRO_PRINCIPAL AS TERCEIRO_PRINCIPAL
	,B.NOVO_TERCEIRO AS NOVO_TERCEIRO
	
 	,(CASE WHEN B.TERCEIRO_PRINCIPAL IS NULL THEN B.NOVO_TERCEIRO
		ELSE B.TERCEIRO_PRINCIPAL END) AS TERCEIRO_AJUSTADO
	,`PARTE_CONTRÁRIA_DATA_ADMISSÃO` AS DATA_ADMISSAO
	,`PARTE_CONTRÁRIA_DATA_DISPENSA` AS DATA_DISPENSA
	,'{mes_fechamento}' AS MES_FECH

	FROM FECHAMENTO_TRAB_2 AS A
	LEFT JOIN TRAB_GER_CONSOLIDA AS B ON A.ID_PROCESSO = B.PROCESSO_ID; 
""")

df_fechamento_trab = convert_to_date_format(df_fechamento_trab, ["MES_FECH"])

# COMMAND ----------

# MAGIC %md
# MAGIC #####Faz o tratamento do campo assunto cargo

# COMMAND ----------

from pyspark.sql.functions import expr

df_fechamento_trab = df_fechamento_trab.withColumn(
    "OBJETO_ASSUNTO_CARGO_M", 
    expr("""
        CASE 
        WHEN INSTR(OBJETO_ASSUNTO_CARGO_M, ' PARA') > 0 THEN 
            SUBSTRING(OBJETO_ASSUNTO_CARGO_M, 1, LENGTH(OBJETO_ASSUNTO_CARGO_M) - 5)
        ELSE 
            OBJETO_ASSUNTO_CARGO_M
        END
    """)
)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Faz o agrupamento dos cargos utilizados no modelo de provisão

# COMMAND ----------

df_fechamento_trab.createOrReplaceTempView(f"TRAB_FECH_CONSOLIDA")

df_fechamento_trab = spark.sql(f"""
/* CARREGA AS INFORMAÇÕES ADICIONAIS NA BASE DO FECHAMENTO */
SELECT *
	,(CASE WHEN OBJETO_ASSUNTO_CARGO_M IN ('AJUDANTE',
										'AJUDANTE EXTERNO',
										'ANALISTA',
										'AUXILIAR',
										'CAIXA',
										'GERENTE',
										'MONTADOR',
										'MOTORISTA',
										'OPERADOR',
										'VENDEDOR') THEN OBJETO_ASSUNTO_CARGO_M
	ELSE 'OUTROS' END) AS CARGO_TRATADO
	FROM TRAB_FECH_CONSOLIDA; 
""")

# COMMAND ----------

# MAGIC %md
# MAGIC #####Conta a quantidade de processos por cargo

# COMMAND ----------

type(df_fechamento_trab)
grouped_df = df_fechamento_trab.groupBy("OBJETO_ASSUNTO_CARGO_M")

# Calculando agregações para cada grupo
agg_df = grouped_df.agg(
    count("OBJETO_ASSUNTO_CARGO_M").alias("count_value")
)

# agg_df.show()
# display(agg_df.orderBy(desc('count_value')))
display(agg_df)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Calcula o tempo de empresa e o aging do estoque

# COMMAND ----------

df_fechamento_trab.createOrReplaceTempView("FECHAMENTO_TRAB_3")
mes_fechamento = dbutils.widgets.get("mes_fechamento")
dt_fech_elaw = dbutils.widgets.get("fech_elaw")

df_fechamento_trab_1 = spark.sql(f"""
/* CARREGA AS INFORMAÇÕES ADICIONAIS NA BASE DO FECHAMENTO */
SELECT *,
    to_date(concat(year(DATA_ADMISSAO), '-', month(DATA_ADMISSAO), '-', '01')) AS MES_DATA_ADMISSAO,
    to_date(concat(year(DATA_DISPENSA), '-', month(DATA_DISPENSA), '-', '01')) AS MES_DATA_DISPENSA,
    year(DATA_DISPENSA) AS ANO_DISPENSA,
    /*((DATA_DISPENSA - DATA_ADMISSAO) / 30) AS TEMPO_EMPRESA_MESES,*/
    datediff(DATA_DISPENSA, DATA_ADMISSAO) AS TEMPO_EMPRESA_MESES,
    (CASE WHEN ESTOQUE = 1 THEN (datediff(to_date('{dt_fech_elaw}'), CADASTRO_AJUSTADO) / 30) END) AS AGING_ESTOQ_MESES,
    (CASE WHEN ESTOQUE = 1 THEN (datediff(to_date('{dt_fech_elaw}'), CADASTRO_AJUSTADO) ) END) AS ANO_AGING_ESTOQ,
    to_date('{dt_fech_elaw}') AS FECH_ELAW

    /*(CASE WHEN ESTOQUE = 1 THEN PUT(({dt_fech_elaw} - CADASTRO_AJUSTADO) / 30, AgingEstoque.) END) AS FX_MES_AGING_ESTOQ,
    (CASE WHEN ESTOQUE = 1 THEN PUT(({dt_fech_elaw} - CADASTRO_AJUSTADO), AgingNovoModelo.) END) AS FX_ANO_AGING_ESTOQ,
   '{mes_fechamento}' AS MES_FECH */
FROM FECHAMENTO_TRAB_3;
""")

# COMMAND ----------

display(df_fechamento_trab_1.select("DATA_DISPENSA", "DATA_ADMISSAO","ANO_DISPENSA", "TEMPO_EMPRESA_MESES","ESTOQUE", "FECH_ELAW","CADASTRO_AJUSTADO","ANO_AGING_ESTOQ"))

# COMMAND ----------

# MAGIC %md
# MAGIC #####Cria as faixas para o cálculo entre as datas

# COMMAND ----------

from pyspark.sql.types import StructType, StructField, IntegerType, StringType

# AgingEstoque
data = [
    (-9999, 6, 'até 6 meses'),
    (6, 12, '7 - 12 meses'),
    (12, 18, '13 - 18 meses'),
    (18, 24, '19 - 24 meses'),
    (24, 9999999, '+24 meses')
]

# Define the schema for the DataFrame
schema = StructType([
    StructField("start_range", IntegerType(), False),
    StructField("end_range", IntegerType(), False),
    StructField("label", StringType(), False)
])

df_AgingEstoque = spark.createDataFrame(data, schema)
df_AgingEstoque.createOrReplaceTempView("AgingEstoque")
df_AgingEstoque.show()

# COMMAND ----------

# AgingTrab
data = [
    (-9999, 360, 'até 1 ano'),
    (360, 1080, '1 - 3 anos'),
    (1080, 2160, '3 - 6 anos'),
    (2160, 3240, '6 - 9 anos'),
    (3240, 99999999, '+9 anos')
]

# Define the schema for the DataFrame
schema = StructType([
    StructField("start_range", IntegerType(), False),
    StructField("end_range", IntegerType(), False),
    StructField("label", StringType(), False)
])

df_AgingTrab = spark.createDataFrame(data, schema)
df_AgingTrab.createOrReplaceTempView("AgingTrab")
df_AgingTrab.show()

# COMMAND ----------

# AgingNovoModelo

data = [
    (-9999, 360, 'ATÉ 1 ANO'),
    (360, 720, '1 - 2 ANOS'),
    (720, 1080, '2 - 3 ANOS'),
    (1080, 1440, '3 - 4 ANOS'),
    (1440, 1800, '4 - 5 ANOS'),
    (1800, 2160, '5 - 6 ANOS'),
    (2160, 2520, '6 - 7 ANOS'),
    (2520, 9999999, '+7 ANOS')
]

# Define the schema for the DataFrame
schema = StructType([
    StructField("start_range", IntegerType(), False),
    StructField("end_range", IntegerType(), False),
    StructField("label", StringType(), False)
])

df_AgingNovoModelo = spark.createDataFrame(data, schema)
df_AgingNovoModelo.createOrReplaceTempView("AgingNovoModelo")
df_AgingNovoModelo.show()

# COMMAND ----------

# AgingTempoEmpresa

data = [
    (-9999, 360, 'Até 1 ano'),
    (360, 720, '1 - 2 anos'),
    (720, 1080, '2 - 3 anos'),
    (1080, 1440, '3 - 4 anos'),
    (1440, 1800, '4 - 5 anos'),
    (1800, 2160, '5 - 6 anos'),
    (2160, 3600, '6 - 10 anos'),
    (3600, 9999999, 'Acima de 10 anos'),
]

# Define the schema for the DataFrame
schema = StructType([
    StructField("start_range", IntegerType(), False),
    StructField("end_range", IntegerType(), False),
    StructField("label", StringType(), False)
])

df_AgingTempoEmpresa = spark.createDataFrame(data, schema)
df_AgingTempoEmpresa.createOrReplaceTempView("AgingTempoEmpresa")
df_AgingTempoEmpresa.show()

# COMMAND ----------

# MAGIC %md
# MAGIC #####Cria as faixas de acordo com as regras de período

# COMMAND ----------

df_fechamento_trab_1.createOrReplaceTempView("FECHAMENTO_TRAB_4")
df_dp_terceiro.createOrReplaceTempView("TB_DP_TERCEIROS")

df_fechamento_trab_2 = spark.sql("""
          SELECT A.*
          ,B.label AS FX_MES_AGING_ESTOQ
          ,C.label AS FX_ANO_AGING_ESTOQ
          ,(CASE WHEN A.TEMPO_EMPRESA_MESES IS NULL THEN 'Sem info' ELSE
                 D.label END) AS `Cluster Aging Tempo de Empresa`

          ,(CASE WHEN A.ANO_AGING_ESTOQ IS NULL THEN 'Sem info' ELSE
                 E.label END) AS `Cluster Aging`

          /*,(datediff(DATA_DISPENSA, DATA_ADMISSAO)/360) AS TEMPO_EMPRESA_ANOS*/
          
          ,(CASE WHEN datediff(DATA_DISPENSA, DATA_ADMISSAO) <= 1800 THEN 
              CAST(((year(DATA_DISPENSA)) - round((datediff(DATA_DISPENSA, DATA_ADMISSAO)/360),0)) AS INT)
                     WHEN datediff(DATA_DISPENSA, DATA_ADMISSAO) > 1800 THEN
              CAST(round(((year(DATA_DISPENSA)) - 5),0) AS INT) ELSE 'Sem info' END) AS `Safra de Reclamação`

          ,(CASE WHEN F.`EMPRESA TERCEIRIZADA` IS NULL THEN 'Sem info' ELSE
                     F.TERCEIRO_AJUSTADO END) AS ET

          FROM 
          FECHAMENTO_TRAB_4 A
          LEFT JOIN AgingEstoque B ON A.AGING_ESTOQ_MESES > B.start_range AND A.AGING_ESTOQ_MESES <= B.end_range
          LEFT JOIN AgingNovoModelo C ON A.ANO_AGING_ESTOQ > C.start_range AND A.ANO_AGING_ESTOQ <= C.end_range
          LEFT JOIN AgingTempoEmpresa D ON A.TEMPO_EMPRESA_MESES > D.start_range AND A.TEMPO_EMPRESA_MESES <= D.end_range
          LEFT JOIN AgingNovoModelo E ON A.ANO_AGING_ESTOQ > E.start_range AND A.ANO_AGING_ESTOQ <= E.end_range
          LEFT JOIN TB_DP_TERCEIROS AS F ON A.TERCEIRO_AJUSTADO = F.`EMPRESA TERCEIRIZADA`;
          """
)

df_fechamento_trab_2 = df_fechamento_trab_2.drop('ANO_DISPENSA')

# COMMAND ----------

# MAGIC %md
# MAGIC #####Cria variável data com o formato ano + mes + dia. exemplo: 20240620

# COMMAND ----------

# nmtabela = nmtabela_trab_ger_consolidado[:6]
dtanomesdia = nmtabela_finan[6:10] + nmtabela_finan[3:5] + nmtabela_finan[0:2]
print(dtanomesdia)

# COMMAND ----------

display(df_fechamento_trab_2)

# COMMAND ----------

# MAGIC %md
# MAGIC #####Renomeia os campos para rodar o modelo de provisão

# COMMAND ----------

df_fechamento_trab_2 = df_fechamento_trab_2 \
    .withColumnRenamed("ID_PROCESSO", "ID PROCESSO") \
    .withColumnRenamed("ÁREA_DO_DIREITO", "Área do Direito")\
    .withColumnRenamed("SUB_ÁREA_DO_DIREITO", "Sub-área do Direito")\
    .withColumnRenamed("OBJETO_ASSUNTO_CARGO_M", "Objeto Assunto/Cargo (M)")\
    .withColumnRenamed("NATUREZA_OPERACIONAL_M", "Natureza Operacional (M)")

# COMMAND ----------

import pandas as pd
from shutil import copyfile

# Convert PySpark DataFrame to Pandas DataFrame
pandas_df = df_fechamento_trab_2.toPandas()

# Save the Pandas DataFrame to an Excel file
local_path = f'/local_disk0/tmp/Trabalhista_Automacao_F.xlsx'
pandas_df.to_excel(local_path, index=False, sheet_name='Automacao', engine='xlsxwriter')

# Copy the file from the local disk to the desired volume
volume_path = f'/Volumes/databox/juridico_comum/arquivos/modelo_provisao/output/FECH_TRAB_MODELAGEM_{dtanomesdia}_F.xlsx'

copyfile(local_path, volume_path)

# COMMAND ----------

# Bloco de código para consultar um id especifico

import pandas as pd

# Read the Excel file into a Pandas DataFrame
file_path = '/Volumes/databox/juridico_comum/arquivos/trabalhista/bases_fechamento_financeiro/trabalhista_base_financeiro_tratada/FECHAMENTO_TRAB_202212.xlsx'
df = pd.read_excel(file_path)

# Convert column ID PROCESSO to integer
df['ID PROCESSO'] = df['ID PROCESSO'].astype('int')

# Filter the DataFrame for the ID 56790
filtered_df = df[df['ID PROCESSO'] == 56790]

# Display the filtered DataFrame
filtered_df

# COMMAND ----------

# MAGIC %sql
# MAGIC -- Lista catalogos
# MAGIC -- SHOW CATALOGS
# MAGIC
# MAGIC -- lista esquemas dentro de um catalogo
# MAGIC -- SHOW SCHEMAS IN databox
# MAGIC
# MAGIC -- Lista tabelas dentro de um esquema (que está dentro de um catalogo) [voce precisa ter permissao de acesso ao catalogo]
# MAGIC -- SHOW TABLES IN databox.juridico_comum
# MAGIC
# MAGIC SELECT COUNT(ID_PROCESSO) AS QTD, SUB_OBJETO_ASSUNTO_CARGO_M
# MAGIC  FROM databox.juridico_comum.tb_fecham_trab_202406
# MAGIC  GROUP BY SUB_OBJETO_ASSUNTO_CARGO_M
# MAGIC  ORDER BY QTD DESC

# COMMAND ----------

nome_tabela = 'tb_fecham_trab_202406'

df = spark.sql(f"SELECT * FROM databox.juridico_comum.{nome_tabela}")

row_count = df.count()

print(f'Dataframe com {row_count} linhas')
display(df)

# COMMAND ----------

import pandas as pd

dt_arquivo_pagamentos = '20240708'

base_pagamentos = f'/Volumes/databox/juridico_comum/arquivos/bases_pagamentos/HISTORICO_DE-PAGAMENTOS_{dt_arquivo_pagamentos}.xlsx'

df_pagamentos = spark.read \
    .format("com.crealytics.spark.excel") \
    .option("header", "true") \
    .option("inferSchema", "true") \
    .option("dataAddress", "PAGAMENTOS!A1") \
    .load(base_pagamentos)

# df_pagamentos = pd.read_excel(base_pagamentos, dtype=str, engine='openpyxl')

print(':D')


# COMMAND ----------

df_pagamentos_colunas = df_pagamentos.columns

if 'PROCESSO - ID' not in df_pagamentos_colunas:
    df_pagamentos_formatado = df_pagamentos.iloc[4:]
    df_pagamentos_formatado = df_pagamentos_formatado.reset_index(drop=True)
    new_header = df_pagamentos_formatado.iloc[0]
    df_pagamentos_formatado = df_pagamentos_formatado.iloc[1:]
    df_pagamentos_formatado.columns = new_header
    df_pagamentos_formatado = df_pagamentos_formatado.reset_index(drop=True)
    
display(df_pagamentos_formatado.head(5))

# COMMAND ----------

from pyspark.sql.functions import col, concat_ws
from collections import Counter
import re
spark.conf.set("spark.sql.caseSensitive","true")
 
dt_arquivo_pagamentos = '20240708'

base_pagamentos = f'/Volumes/databox/juridico_comum/arquivos/bases_pagamentos/HISTORICO_DE-PAGAMENTOS_{dt_arquivo_pagamentos}.xlsx'

# Importação da Base do Calculista de Pedidos
df_calculista_processos = spark.read.format("com.crealytics.spark.excel") \
    .option("header", "true") \
    .option("inferSchema", "true") \
    .option("ignoreLeadingWhiteSpace", "true") \
    .option("treatEmptyValuesAsNulls", "true") \
    .option("ignoreTrailingWhiteSpace", "true") \
    .option("MaxRowsInMemory", 1000) \
    .option("dataAddress", "'PAGAMENTOS'!A6:AM1048576") \
    .load(base_pagamentos)
 
# Get the column names
columns = df_calculista_processos.columns
 
# Count the occurrences of each column name
column_counts = Counter(columns)
 
# Create a list to store unique column names
unique_columns = []
 
# Iterate over column names and their counts
for column, count in column_counts.items():
    if count > 1:
        # If the column name appears more than once, rename the duplicates
        for i in range(count):
            # Replace spaces and special characters with underscores
            cleaned_column = re.sub(r'[\s\W]', '_', column)
            old_name = concat_ws("_", cleaned_column, str(i))
            new_name = concat_ws("_", cleaned_column, str(i+1)) if i > 0 else cleaned_column
            df_calculista_processos = df_calculista_processos.withColumnRenamed(old_name, new_name)
    else:
        # If the column name is unique, add it to unique_columns
        unique_columns.append(column)

display(df_calculista_processos)

# COMMAND ----------

tamanho = df_calculista_processos.count()

print(tamanho)
